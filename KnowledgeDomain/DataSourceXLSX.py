__author__ = 'Jonathan Langens'
from openpyxl import load_workbook
from openpyxl.cell import Cell
from TemplateXLSX import TemplateXLSX
from DataSource import DataSource

class DataSourceXLSX(DataSource):
    def __init__(self, filename):
        self.template = TemplateXLSX()
        self.sourceFile = "/home/jeuna/Downloads/datasources/wb.xlsx"

    def addURIIOS(self, instance, URIIOCriteria, domain):
        # read in the source file
        wb = load_workbook(self.sourceFile)
        ws = wb.get_active_sheet()

        # we iterate over all templates and try to read the data that they contain
        eligilbeObjectTemplates = self.template.getObjectTemplatesForCriteria(URIIOCriteria, domain)

        for ot in eligilbeObjectTemplates:
            uriio = instance.uriioManager.newURIIO()
            utype = domain.typeManager.getType(ot.type)
            if utype is None:
                utype = domain.typeManager.getType("type")
            uriio.addType(utype)

            for p in ot.properties:
                pname = p.name
                if pname == "" or pname is None:
                    pname = str(self.getDataForPosition(ws, p.namex, p.namey))
                pvalue = self.getDataForPosition(ws, p.x, p.y)
                ptype = self.getTypeForPosition(ws, p.x, p.y)
                uriio.addProperty(pname, pvalue, ptype)

        return True

    def addPredicates(self, instance, predicateCriteria, domain):
        # read in the source file
        wb = load_workbook(self.sourceFile)
        ws = wb.get_active_sheet()

        # we iterate over all templates and try to read the data that they contain
        eligilbeObjectTemplates = self.template.getObjectTemplatesForCriteria(predicateCriteria, domain)

        for ot in eligilbeObjectTemplates:
            for conn in ot.connectors:
                uc = instance.uriioManager.getCriteria()
                for p in ot.properties:
                    pname = p.name
                    if pname == "" or pname is None:
                        pname = str(self.getDataForPosition(ws, p.namex, p.namey))
                    pvalue = self.getDataForPosition(ws, p.x, p.y)
                    uc.addPropertyRestriction(pname, pvalue)
                urilist = uc.getURIIOS()
                if len(urilist) > 0:
                    otURI = urilist[0].URI

                    connValue = self.getDataForPosition(ws, conn.x, conn.y)
                    uriioCriteria = instance.uriioManager.getCriteria()
                    uriioCriteria.TypeRestrictions.append(conn.type)
                    uriioCriteria.addPropertyRestriction(conn.property, connValue)
                    uriioCriteria.resolve()
                    for uriio in uriioCriteria.getURIIOs():
                        instance.predicateManager.addPredicate(uriio.URI, conn.predicate, otURI)


    def getTypeForPosition(self, data, rowpos, colpos):
        rowpos = int(rowpos)
        colpos = int(colpos)
        if rowpos < 1 or colpos < 1:
            return "text"
        if data is not None:
            dt = data.cell(row=rowpos, column=colpos).data_type
            if dt == Cell.TYPE_STRING:
                return "text"
            elif dt ==  Cell.TYPE_NUMERIC:
                nf = data.cell(row=rowpos, column=colpos).number_format
                if nf == 'dd/yy':
                    return "date"
                elif nf == 'yyyy-mm-dd':
                    return "date"
                elif nf == 'yyyy-mm-dd h:mm:ss':
                    return "date"
                elif nf == 'MM/DD/YY':
                    return "date"
                elif nf == "GENERAL":
                    return "number"
                return "number"
        return "text"


    def getDataForPosition(self, data, rowpos, colpos):
        rowpos = int(rowpos)
        colpos = int(colpos)
        if rowpos < 1 or colpos < 1:
            return ""

        if data is not None:
            return data.cell(row=rowpos, column=colpos).value

        return ""