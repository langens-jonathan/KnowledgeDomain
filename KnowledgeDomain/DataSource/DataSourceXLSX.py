__author__ = 'Jonathan Langens'
from openpyxl import load_workbook
from openpyxl.cell import Cell
from KnowledgeDomain.Template.TemplateXLSX import TemplateXLSX
from KnowledgeDomain.DataSource.DataSource import DataSource
from KnowledgeDomain.Template.TemplateXLSX import XLSXObjectConnectorTemplate
from KnowledgeDomain.Template.TemplateXLSX import XLSXObjectExtenderConnectorTemplate

class DataSourceXLSX(DataSource):
    def __init__(self, filename):
        self.template = TemplateXLSX()
        self.sourceFile = "/home/jeuna/Downloads/datasources/wb.xlsx"

    def processPreQueries(self, userBox, criteria, domain):
        return False

    def processConnectors(self, userBox, criteria, domain):
        # read in the source file
        wb = load_workbook(self.sourceFile)
        ws = wb.get_active_sheet()

        for ot in self.template.objectTemplates:
            print("*** DETECTING connectors for: " + str(self.getDataForPosition(ws, ot.properties[0].x, ot.properties[0].y)))
            # first get the object whom this ot belongs to out of the knowledge instance
            tcrit = userBox.knowledgeInstance.uriioManager.getCriteria()
            print("*** we start with " + str(len(tcrit.URIIOList)) + " uriios")
            for op in ot.properties:
                pname = op.name
                if pname == "" or pname is None:
                    pname = str(self.getDataForPosition(ws, op.namex, op.namey))
                pvalue = str(self.getDataForPosition(ws, op.x, op.y))
                if pname == "id":
                    tcrit.addPropertyRestriction(pname, pvalue)
                    print("*** adding property " + pname + " with value: " + pvalue)
                    tcrit.addPropertyRestriction("id-x-location", str(op.x))
                    tcrit.addPropertyRestriction("id-y-location", str(op.y))
            tcrit.resolve()
            print("*** we end with " + str(len(tcrit.URIIOList)) + " uriios")
            if len(tcrit.URIIOList) > 0:
                print("***WE FOUND OBJET WITH URIIO: " + tcrit.URIIOList[0].URI + " ***")
                # now doing this stuff for every predicate connector
                for pc in ot.predicateConnectors:
                    # first create a new criteria object
                    crit = userBox.knowledgeInstance.uriioManager.getCriteria()
                    tp = domain.typeManager.getType(pc.type)
                    if not tp is None:
                        crit.TypeRestrictions.append(tp)
                    crit.addPropertyRestriction(pc.property, self.getDataForPosition(ws, pc.x, pc.y))

                    # next resolve it
                    crit.resolve()

                    # now get the predicate
                    pred = domain.predicateDefinitionManager.getPredicate(pc.predicate)
                    if pred is None:
                        pred = domain.predicateDefinitionManager.getPredicate("refersTo")

                    # then add a predicate to the predicate manager for every uriio that survived the criteria
                    print("the number of eligible uriios is " + str(len(crit.URIIOList)))
                    for u in crit.URIIOList:
                        print("adding connection with " + u.asXML())
                        userBox.knowledgeInstance.predicateManager.addPredicate(u, pred, tcrit.URIIOList[0])
            else:
                print("*** NOT FOUND")
                if tcrit is not None:
                    print("*** criteria used: " + tcrit.__str__())
                else:
                    print("*** criteria is none...")

        return True

    def extractURIIOs(self, userBox, criteria, domain):
        # read in the source file
        wb = load_workbook(self.sourceFile)
        ws = wb.get_active_sheet()

        # setting the instance
        instance = userBox.knowledgeInstance

        # we iterate over all templates and try to read the data that they contain
        eligilbeObjectTemplates = self.template.getObjectTemplatesForCriteria(criteria, domain)

        for ot in eligilbeObjectTemplates:
            uriio = instance.uriioManager.newURIIO()
            utype = domain.typeManager.getType(ot.type)
            if utype is None:
                utype = domain.typeManager.getType("type")
            uriio.addType(utype)

            for p in ot.properties:
                pname = p.name
                if pname == "" or pname is None:
                    pname = str(self.getDataForPosition(ws, p.namex, p.namey))
                pvalue = str(self.getDataForPosition(ws, p.x, p.y))
                ptype = str(self.getTypeForPosition(ws, p.x, p.y))
                uriio.addProperty(pname, pvalue, ptype)
                if pname == "id":
                    uriio.addProperty("id-x-location", str(p.x), "text")
                    uriio.addProperty("id-y-location", str(p.y), "text")

            for pt in ot.predicateConnectors:
                pname = "predicateConnector"
                pvalue = str(self.getDataForPosition(ws, pt.x, pt.y))
                uriio.addProperty(pname, pvalue, "text")

        return True


    """
    add predicates will change as predicates will be added through the connectors
    """
    def addPredicates(self, instance, predicateCriteria, domain):
        # read in the source file
        wb = load_workbook(self.sourceFile)
        ws = wb.get_active_sheet()

        # we iterate over all templates and try to read the data that they contain
        eligilbeObjectTemplates = self.template.getObjectTemplatesForCriteria(predicateCriteria, domain)

        for ot in eligilbeObjectTemplates:
            for conn in ot.connectors:
                uc = instance.uriioManager.getCriteria()
                for p in ot.properties:
                    pname = p.name
                    if pname == "" or pname is None:
                        pname = str(self.getDataForPosition(ws, p.namex, p.namey))
                    pvalue = self.getDataForPosition(ws, p.x, p.y)
                    uc.addPropertyRestriction(pname, pvalue)
                urilist = uc.getURIIOS()
                if len(urilist) > 0:
                    otURI = urilist[0].URI

                    connValue = self.getDataForPosition(ws, conn.x, conn.y)
                    uriioCriteria = instance.uriioManager.getCriteria()
                    uriioCriteria.TypeRestrictions.append(conn.type)
                    uriioCriteria.addPropertyRestriction(conn.property, connValue)
                    uriioCriteria.resolve()
                    for uriio in uriioCriteria.getURIIOs():
                        instance.predicateManager.addPredicate(uriio.URI, conn.predicate, otURI)

    def getTypeForPosition(self, data, rowpos, colpos):
        rowpos = int(rowpos)
        colpos = int(colpos)
        if rowpos < 1 or colpos < 1:
            return "text"
        if data is not None:
            dt = data.cell(row=rowpos, column=colpos).data_type
            if dt == Cell.TYPE_STRING:
                return "text"
            elif dt ==  Cell.TYPE_NUMERIC:
                nf = data.cell(row=rowpos, column=colpos).number_format
                if nf == 'dd/yy':
                    return "date"
                elif nf == 'yyyy-mm-dd':
                    return "date"
                elif nf == 'yyyy-mm-dd h:mm:ss':
                    return "date"
                elif nf == 'MM/DD/YY':
                    return "date"
                elif nf == "GENERAL":
                    return "number"
                return "number"
        return "text"


    def getDataForPosition(self, data, rowpos, colpos):
        rowpos = int(rowpos)
        colpos = int(colpos)
        if rowpos < 1 or colpos < 1:
            return ""

        if data is not None:
            return data.cell(row=rowpos, column=colpos).value

        return ""